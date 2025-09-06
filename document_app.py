import os
import sys
import logging
import json
import re
import io
import time
import zipfile
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
from datetime import datetime

# --- Core Libraries ---
import streamlit as st
from PIL import Image
import google.generativeai as genai
import pandas as pd

# --- PDF Handling ---
try:
    import fitz  # PyMuPDF
except ImportError:
    st.error("PyMuPDF is not installed. Please run 'pip install PyMuPDF' to handle PDF files.")
    sys.exit(1)

# --- Logging ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - [%(levelname)s] - %(message)s')

# --- Document Processing Configuration ---
@dataclass
class DocumentProcessorConfig:
    """Configuration for the advanced document processing agent."""
    GOOGLE_API_KEY: str = st.secrets.get("GOOGLE_API_KEY", "")
    GEMINI_MODEL: str = "gemini-1.5-pro"
    OUTPUT_DIR: str = "extracted_documents"

class AdvancedDocumentProcessor:
    """
    Ultra-precise document processing engine that extracts EVERY SINGLE DETAIL from documents
    using forensic-level analysis with ZERO information loss. This is the DEFINITIVE document
    extraction system that leaves no stone unturned.
    """
    
    def __init__(self):
        self.config = DocumentProcessorConfig()
        if not self.config.GOOGLE_API_KEY:
            st.error("Your GOOGLE_API_KEY is not set! Please add it to your configuration.")
            st.stop()
        self._initialize_gemini()
        os.makedirs(self.config.OUTPUT_DIR, exist_ok=True)

    def _initialize_gemini(self):
        try:
            genai.configure(api_key=self.config.GOOGLE_API_KEY)
            self.model = genai.GenerativeModel(self.config.GEMINI_MODEL)
            logging.info("Gemini initialized successfully for advanced document processing.")
        except Exception as e:
            st.error(f"Failed to initialize Gemini. Error: {e}")
            raise

    def extract_document_with_forensic_precision(self, img: Image.Image, doc_type: str, custom_fields: List[str] = None) -> Dict:
        """Extract from single document/page."""
        return self.extract_documents_batch([img], doc_type, custom_fields)[0]
    
    def extract_documents_batch(self, images: List[Image.Image], doc_type: str, custom_fields: List[str] = None) -> List[Dict]:
        """
        BATCH document extraction - processes multiple images in a single API call for cost efficiency.
        Since Gemini charges per TOKEN not per REQUEST, batching saves massive costs on prompt tokens.
        """
        try:
            # Modify prompt for batch processing
            master_prompt = self._construct_forensic_extraction_prompt(doc_type, custom_fields, batch_mode=len(images) > 1)
            
            # Create content list with prompt + all images
            content = [master_prompt] + images
            
            response = self.model.generate_content(content)
            cleaned_text = response.text.strip()
            
            # Aggressive cleaning to ensure valid JSON
            cleaned_text = cleaned_text.replace("```json", "").replace("```", "")
            # Remove any comment patterns that might slip through
            import re
            cleaned_text = re.sub(r'//.*?(?=\n|$)', '', cleaned_text, flags=re.MULTILINE)  # Remove // comments
            cleaned_text = re.sub(r'/\*.*?\*/', '', cleaned_text, flags=re.DOTALL)  # Remove /* */ comments  
            cleaned_text = cleaned_text.strip()
            
            if not cleaned_text:
                st.warning("The Document Analyst (Gemini) returned an empty response.")
                return [{}] * len(images)
            
            # For batch processing, expect array of results or single result to replicate
            if len(images) > 1:
                # Try to parse as array of results
                try:
                    if cleaned_text.startswith('['):
                        results_array = json.loads(cleaned_text)
                        if isinstance(results_array, list) and len(results_array) == len(images):
                            return results_array
                except:
                    pass
                    
                # If not array format, treat as single result and replicate
                st.warning(f"Expected {len(images)} results but got single result - replicating for all images")
            
            # Handle single result (or fallback for batch)
            # Ensure response starts and ends with proper JSON braces
            if not cleaned_text.startswith('{'):
                # Find the first { and start from there
                start_idx = cleaned_text.find('{')
                if start_idx != -1:
                    cleaned_text = cleaned_text[start_idx:]
                else:
                    st.error("No valid JSON object found in response")
                    return [{}] * len(images)
            
            if not cleaned_text.endswith('}'):
                # Find the last } and end there
                end_idx = cleaned_text.rfind('}')
                if end_idx != -1:
                    cleaned_text = cleaned_text[:end_idx + 1]
                else:
                    st.error("No valid JSON object termination found in response")
                    return [{}] * len(images)
            
            # Direct JSON parsing - Gemini should output clean JSON
            try:
                single_result = json.loads(cleaned_text)
            except json.JSONDecodeError as je:
                st.error(f"JSON parsing failed: {je}")
                # Try to fix truncated JSON
                try:
                    # Count open and close braces to detect truncation
                    open_braces = cleaned_text.count('{')
                    close_braces = cleaned_text.count('}')
                    
                    if open_braces > close_braces:
                        # JSON is truncated, try to close it properly
                        missing_closes = open_braces - close_braces
                        fixed_text = cleaned_text + '}' * missing_closes
                        
                        # Also fix any incomplete string values
                        if cleaned_text.rstrip().endswith('"'):
                            # Complete string was cut off, add closing quote
                            fixed_text = cleaned_text.rstrip()[:-1] + '""' + '}' * missing_closes
                        elif not cleaned_text.rstrip().endswith(('}', ']', '"')):
                            # Value was cut off mid-way, close the string and object
                            fixed_text = cleaned_text.rstrip() + '"}' + '}' * (missing_closes - 1)
                        
                        single_result = json.loads(fixed_text)
                        st.warning("JSON was truncated but auto-fixed successfully")
                    else:
                        raise je
                except:
                    st.text_area("Raw response for debugging:", cleaned_text[:2000], height=300)
                    st.error("Could not auto-fix truncated JSON. Response may be too long for API.")
                    return [{}] * len(images)
            
            # Validate extraction completeness
            field_count = len(single_result.get('extracted_fields', {}))
            confidence = single_result.get('extraction_confidence', 0)
            
            # Ensure confidence is a number
            try:
                confidence = float(confidence) if confidence is not None else 0.0
            except (ValueError, TypeError):
                confidence = 0.0
            
            
            # Return single result or replicate for all images
            return [single_result] if len(images) == 1 else [single_result] * len(images)
            
        except Exception as e:
            st.error(f"document extraction failed: {e}")
            logging.error(f"Extraction error for {doc_type}. Response: {response.text if 'response' in locals() else 'N/A'}")
            return [{}] * len(images)

    def _construct_forensic_extraction_prompt(self, doc_type: str, custom_fields: List[str] = None, batch_mode: bool = False) -> str:
        """
        Construct the ULTIMATE document analysis prompt that leaves NO DETAIL unexplored.
        This is based on the same methodology as the table extraction prompt but adapted for documents.
        """
        
        base_forensic_protocol = f"""
        You are an ELITE DOCUMENT INTELLIGENCE ARCHITECT - the world's most advanced forensic document analyst with superhuman pattern recognition and structural understanding. You possess decades of specialized expertise in analyzing complex official documents with ABSOLUTE ZERO TOLERANCE for missing information.

        MISSION: Perform a REVOLUTIONARY FORENSIC ANALYSIS that goes beyond simple OCR. You must ARCHITECT a complete understanding of this document's information ecosystem - its visual hierarchy, semantic relationships, field interdependencies, and contextual meaning.

        CRITICAL UNDERSTANDING: You are NOT following a template or skeleton. You are a MASTER DETECTIVE who must:
        - DISCOVER the document's natural structure
        - BUILD the information architecture from scratch
        - UNDERSTAND the document's unique layout language
        - EXTRACT meaning from spatial relationships and visual cues
        
        🚨 ABSOLUTE RULE: ONLY EXTRACT FIELDS THAT ACTUALLY EXIST ON THE DOCUMENT 🚨
        - DO NOT create fields with "None", "N/A", or empty values
        - DO NOT invent fields based on document type expectations
        - ONLY include fields you can physically see with text/values on the document
        - If a field doesn't exist, don't include it in the extracted_fields at all

        YOUR REVOLUTIONARY ANALYSIS METHODOLOGY:

        ╔══════════════════════════════════════════════════════════════════════════════════════╗
        ║ PHASE 1: FORENSIC DOCUMENT ARCHAEOLOGY & STRUCTURAL DECODING                        ║
        ╚══════════════════════════════════════════════════════════════════════════════════════╝
        
        🔍 DOCUMENT AUTOPSY:
           - Perform microscopic visual forensics of EVERY pixel
           - Map the document's NATURAL information architecture
           - Identify visual hierarchy patterns, groupings, and relationships
           - Decode the document's intrinsic organizational language
           - Understand how information flows spatially across the document
           - Recognize official formatting conventions and regulatory structures

        🧬 LAYOUT DNA ANALYSIS:
           - Analyze spacing patterns, alignment systems, and visual relationships
           - Understand how fields relate to each other spatially and semantically
           - Decode the document's inherent categorization system
           - Map information clusters and logical groupings
           - Identify primary, secondary, and tertiary information levels

        ╔══════════════════════════════════════════════════════════════════════════════════════╗
        ║ PHASE 2: QUANTUM TEXT EXTRACTION & LINGUISTIC INTELLIGENCE                         ║
        ╚══════════════════════════════════════════════════════════════════════════════════════╝

        🚀 HYPER-PRECISE OCR INTELLIGENCE WITH SPATIAL MAPPING:
           - Extract EVERY SINGLE VISIBLE CHARACTER with forensic precision
           - Process ALL languages present WITHOUT translation or conversion
           - PRESERVE the exact multilingual structure as it appears on the document
           - Capture faded, partial, or damaged text using advanced pattern recognition
           - Extract text from ALL elements: headers, labels, values, stamps, watermarks
           - Process handwritten text with calligraphic analysis
           - Preserve EXACT formatting, positioning, and visual context
           - NEVER translate, transliterate, or convert languages - keep original scripts
           - CRITICAL: Record EXACT SPATIAL COORDINATES for every text element
           - Map VERTICAL and HORIZONTAL positioning of all fields
           - Capture RELATIVE positioning between elements
           - Document ALIGNMENT patterns (left, center, right, justified)
           - Record SPACING measurements between elements
           - Identify VISUAL GROUPINGS and their spatial boundaries

        🔗 SEMANTIC RELATIONSHIP MAPPING:
           - Understand the MEANING behind spatial positioning
           - Map label-to-value relationships even when not visually obvious
           - Recognize implied fields and calculated values
           - Understand contextual dependencies between different sections
           - Decode abbreviations, codes, and technical terminology

        ╔══════════════════════════════════════════════════════════════════════════════════════╗
        ║ PHASE 3: SPATIAL ARCHITECTURE & VISUAL LAYOUT RECONSTRUCTION                       ║
        ╚══════════════════════════════════════════════════════════════════════════════════════╝

        🏗️ DOCUMENT SPATIAL MAPPING - CAPTURE THE EXACT VISUAL LAYOUT:
           - Create a COORDINATE SYSTEM for the entire document (0,0 = top-left)
           - Map EVERY text element to its exact X,Y position
           - Record WIDTH and HEIGHT of text blocks
           - Identify ROWS and COLUMNS in the document layout
           - Capture ALIGNMENT patterns (left-aligned, centered, right-aligned)
           - Document SPACING between elements (margins, padding, gaps)
           - Map VISUAL HIERARCHY (headers, subheaders, body text, captions)
           - Identify TEXT ZONES and their boundaries
           - Record FONT SIZES and text styling differences
           - Capture BACKGROUND COLORS and visual separators
           - Map GROUPED ELEMENTS that belong together visually
           - Document READING ORDER and visual flow patterns

        🎯 LAYOUT RECONSTRUCTION INTELLIGENCE:
           - Understand how to RECREATE the document layout in Excel
           - Map text positions to Excel cell coordinates
           - Preserve VERTICAL alignment and spacing
           - Maintain HORIZONTAL relationships between elements
           - Group related fields into logical Excel sections
           - Preserve the document's visual hierarchy in spreadsheet format

        🎯 CONTEXTUAL INTELLIGENCE:
           - Apply specialized knowledge for document type: {doc_type}
           - Understand regulatory, legal, and cultural contexts
           - Recognize standard formats and official conventions
           - Apply domain expertise (administrative, legal, medical, financial)
           - Decode technical codes, reference numbers, and official terminology

        ╔══════════════════════════════════════════════════════════════════════════════════════╗
        ║ PHASE 4: MULTI-DIMENSIONAL VALIDATION & QUALITY ASSURANCE                          ║
        ╚══════════════════════════════════════════════════════════════════════════════════════╝

        🔬 FORENSIC VALIDATION:
           - Cross-validate ALL extracted data for internal consistency
           - Verify calculated fields, checksums, and derived values
           - Validate date formats, numerical consistency, and logical relationships
           - Check against expected document standards and conventions
           - Flag inconsistencies, anomalies, or potential data quality issues

        🎖️ CONFIDENCE ARCHITECTURE:
           - Provide granular confidence scores for each extraction element
           - Assess OCR quality, spatial analysis precision, and semantic accuracy
           - Evaluate overall extraction completeness and reliability
           - Identify elements requiring human verification or validation
           - Rate the quality of source document and extraction challenges

        CRITICAL LAYOUT ANALYSIS - UNDERSTAND THE DOCUMENT'S SPATIAL ORGANIZATION:
        
        🔍 IDENTIFY THE LAYOUT PATTERN:
        - Look at how fields are arranged spatially on the page
        - Is it a simple 2-column format? (Label | Value)  
        - Is it a 3-column format? (Arabic Label | VALUE | French Label)
        - Are values centered with labels on sides?
        - Are there sections, boxes, or grouped areas?
        - What is the NATURAL reading flow of this document?
        
        📍 COMMON PATTERNS TO RECOGNIZE:
        - Moroccan/Bilingual: Arabic text --- VALUE --- French text (same meaning, different languages)
        - Standard forms: Label: Value pairs
        - Tabular: Organized in rows and columns
        - Sectioned: Different areas for different types of information
        - Mixed: Combination of different layout styles
        
        CRITICAL: BE A PERFECT PHOTOCOPY MACHINE - EXTRACT EXACTLY WHAT YOU SEE:
        
        ⚡ NEVER INVENT OR GUESS DATA - if you don't see it clearly, don't extract it
        ⚡ NEVER TRANSLATE OR CONVERT - if Arabic and French appear together, keep them together exactly as shown
        ⚡ NEVER REORGANIZE - if languages are mixed on same line, extract them mixed
        ⚡ NEVER SEPARATE what appears together - preserve the EXACT spatial arrangement
        ⚡ NEVER ADD MISSING DATA - if a field is empty/unclear, mark it as empty
        ⚡ NEVER STANDARDIZE FORMATS - extract dates, numbers, text exactly as written
        ⚡ COPY THE DOCUMENT LAYOUT EXACTLY - don't impose your own structure
        ⚡ EXTRACT ONLY WHAT EXISTS - don't fill in blanks with assumptions
        ⚡ PRESERVE EVERY CHARACTER - including spaces, punctuation, formatting
        ⚡ BE PIXEL-PERFECT FAITHFUL - like scanning the document into digital form
        """

        # Document-specific extraction protocols
        if doc_type == "identity_card":
            specific_protocol = """
            IDENTITY DOCUMENT EXTRACTION PROTOCOL:

            **PRIMARY IDENTITY FIELDS (EXTRACT IF VISIBLE)**:
            - full_name: Complete legal name exactly as written in ALL languages/scripts present
            - date_of_birth: Birth date in original format + ISO format
            - place_of_birth: Complete birth location details
            - document_number: Primary identification number
            - document_type: Specific type of ID (national ID, passport, etc.)
            - issue_date: Document issuance date
            - expiry_date: Document expiration date
            - issuing_authority: Government body that issued document
            - nationality: Citizenship information
            - gender: Gender/sex designation
            - address: Complete address information in ALL languages exactly as written

            **SECONDARY IDENTITY FIELDS (EXTRACT ONLY IF PRESENT)**:
            - father_name: Paternal name if listed
            - mother_name: Maternal name if listed
            - spouse_name: Spouse information if present
            - profession: Occupation if listed
            - height: Physical height if present
            - eye_color: Eye color if present
            - blood_type: Blood type if present
            - phone_number: Contact number if present
            - emergency_contact: Emergency contact if present

            **DOCUMENT-SPECIFIC FIELD EXTRACTION**:
            - Extract ALL visible labels and their corresponding values exactly as written
            - Preserve original language text for ALL field names and values
            - Identify official numbers, codes, and reference identifiers
            - Process ALL language sections maintaining their distinct identity
            - Capture government/authority names in their original script

            **DOCUMENT SECURITY FEATURES**:
            - photo_present: Boolean indicating photo presence
            - photo_quality: Assessment of photo quality
            - signatures_present: Count and description of signatures
            - stamps_present: Official stamps and seals
            - watermarks: Security watermarks if visible
            - holographic_elements: Security holograms
            - raised_text: Embossed or raised text elements
            - microprint: Micro-printed security text
            - security_features: Other security elements

            **DOCUMENT CONDITION ASSESSMENT**:
            - document_condition: Overall physical condition
            - text_clarity: Clarity of text elements
            - image_quality: Quality of photo if present
            - wear_patterns: Signs of wear or damage
            - authenticity_indicators: Security feature verification
            """

        elif doc_type == "invoice":
            specific_protocol = """
            INVOICE/RECEIPT EXTRACTION PROTOCOL:

            **HEADER INFORMATION (MANDATORY)**:
            - invoice_number: Unique invoice identifier
            - invoice_date: Date of invoice issuance
            - due_date: Payment due date if specified
            - purchase_order_number: PO reference if present
            - invoice_type: Type of invoice (standard, credit, debit)

            **VENDOR INFORMATION (COMPLETE EXTRACTION)**:
            - vendor_name: Complete business name
            - vendor_address: Full business address
            - vendor_address_line_1: Primary address line
            - vendor_address_line_2: Secondary address line
            - vendor_city: City location
            - vendor_state_province: State or province
            - vendor_postal_code: Postal/ZIP code
            - vendor_country: Country
            - vendor_phone: Contact phone number
            - vendor_fax: Fax number if present
            - vendor_email: Email address if present
            - vendor_website: Website if present
            - vendor_tax_id: Tax identification number
            - vendor_business_registration: Business registration number
            - vendor_vat_number: VAT registration number

            **CUSTOMER INFORMATION (COMPLETE EXTRACTION)**:
            - customer_name: Customer/buyer name
            - customer_address: Complete customer address
            - customer_contact: Customer contact information
            - customer_tax_id: Customer tax ID if present
            - billing_address: Billing address if different
            - shipping_address: Shipping address if different

            **LINE ITEMS (EXHAUSTIVE EXTRACTION)**:
            - line_items: Array of all products/services with:
              * item_number: Product/service code
              * description: Complete item description
              * quantity: Quantity ordered/delivered
              * unit_of_measure: Unit measurement (pcs, kg, hours)
              * unit_price: Price per unit
              * discount_rate: Discount percentage if applied
              * discount_amount: Discount amount if applied
              * line_total: Total for this line item
              * tax_rate: Tax rate applied to item
              * tax_amount: Tax amount for item
              * category: Product/service category

            **FINANCIAL CALCULATIONS (PRECISE EXTRACTION)**:
            - subtotal: Pre-tax total amount
            - discount_total: Total discounts applied
            - tax_breakdown: Detailed tax calculations by rate
            - total_tax_amount: Sum of all taxes
            - shipping_cost: Shipping charges if present
            - handling_fee: Handling charges if present
            - other_fees: Any additional fees
            - total_amount: Final amount due
            - amount_paid: Amount already paid if shown
            - balance_due: Outstanding balance
            - currency: Currency used
            - exchange_rate: Currency exchange rate if applicable

            **PAYMENT INFORMATION**:
            - payment_terms: Payment terms and conditions
            - payment_method: Accepted payment methods
            - bank_details: Banking information for payment
            - late_fee_policy: Late payment charges
            - early_payment_discount: Early payment terms

            **ADDITIONAL INVOICE ELEMENTS**:
            - notes: Any additional notes or comments
            - terms_conditions: Terms and conditions
            - return_policy: Return/refund policy
            - warranty_information: Warranty details
            - delivery_information: Shipping/delivery details
            """

        elif doc_type == "medical":
            specific_protocol = """
            MEDICAL DOCUMENT EXTRACTION PROTOCOL:

            **PATIENT INFORMATION (MANDATORY)**:
            - patient_name: Full patient name
            - patient_id: Medical record number
            - date_of_birth: Patient birth date
            - age: Current age if stated
            - gender: Patient gender
            - address: Patient address
            - phone: Patient contact number
            - emergency_contact: Emergency contact information
            - insurance_id: Insurance identification
            - insurance_provider: Insurance company name

            **MEDICAL ENCOUNTER DETAILS**:
            - document_date: Date of medical document
            - visit_date: Date of medical visit
            - admission_date: Hospital admission date if applicable
            - discharge_date: Discharge date if applicable
            - document_type: Type of medical document
            - facility_name: Medical facility name
            - facility_address: Facility address
            - department: Medical department or unit

            **MEDICAL STAFF INFORMATION**:
            - primary_doctor: Attending physician name
            - doctor_license: Medical license number
            - consulting_doctors: Other doctors involved
            - nurse_staff: Nursing staff if mentioned
            - technician: Lab technician if applicable

            **MEDICAL DATA & RESULTS**:
            - chief_complaint: Primary reason for visit
            - symptoms: Reported symptoms
            - vital_signs: Temperature, BP, pulse, etc.
            - test_results: Laboratory test results
            - test_values: Specific numeric values
            - reference_ranges: Normal value ranges
            - abnormal_results: Flagged abnormal values
            - imaging_results: X-ray, MRI, CT scan results
            - pathology_results: Pathology findings

            **DIAGNOSES & TREATMENTS**:
            - primary_diagnosis: Main diagnosis
            - secondary_diagnoses: Additional diagnoses
            - icd_codes: ICD diagnostic codes
            - medications_prescribed: Prescribed medications
            - dosage_instructions: Medication dosages
            - treatment_plan: Recommended treatments
            - procedures_performed: Medical procedures
            - follow_up_instructions: Next steps for patient

            **ADMINISTRATIVE INFORMATION**:
            - medical_record_number: MRN
            - visit_id: Visit identifier
            - billing_codes: Medical billing codes
            - authorization_numbers: Insurance authorization
            - provider_signature: Doctor's signature
            - document_verification: Authentication elements
            """

        elif doc_type == "certificate":
            specific_protocol = """
            CERTIFICATE/DIPLOMA EXTRACTION PROTOCOL:

            **CERTIFICATE IDENTIFICATION**:
            - certificate_title: Full title of certificate
            - certificate_type: Type (diploma, license, award, etc.)
            - certificate_level: Academic/professional level
            - certificate_number: Unique certificate identifier
            - series_number: Series or batch number
            - registration_number: Official registration number

            **RECIPIENT INFORMATION**:
            - recipient_name: Full name of recipient
            - recipient_id: Student/professional ID number
            - date_of_birth: Recipient birth date if present
            - nationality: Recipient nationality if stated

            **ACHIEVEMENT DETAILS**:
            - field_of_study: Subject area or specialization
            - program_name: Specific program completed
            - degree_classification: Honors, grades, classification
            - gpa_score: Grade point average if present
            - completion_date: Program completion date
            - duration_of_study: Length of program
            - thesis_title: Dissertation/thesis title if applicable

            **ISSUING INSTITUTION**:
            - institution_name: Complete institution name
            - institution_address: Institution address
            - institution_type: University, college, organization type
            - accreditation_body: Accrediting organization
            - faculty_department: Specific faculty or department
            - dean_name: Dean or department head name
            - registrar_name: Registrar or official name

            **ISSUANCE DETAILS**:
            - issue_date: Date certificate was issued
            - conferral_date: Date degree/certification was conferred
            - graduation_date: Graduation ceremony date
            - academic_year: Academic year of completion
            - ceremony_location: Graduation ceremony location

            **AUTHENTICATION ELEMENTS**:
            - official_seal: Description of official seals
            - signatures: Names and titles of signatories
            - watermarks: Security watermarks
            - embossed_elements: Raised or embossed features
            - security_features: Other authentication elements
            - verification_code: QR codes or verification numbers
            """

        else:  # Generic form or unknown document
            specific_protocol = """
            GENERIC DOCUMENT EXTRACTION PROTOCOL:

            **DOCUMENT STRUCTURE ANALYSIS**:
            - document_title: Main document title or header
            - document_type: Inferred document type
            - form_number: Form number or identifier
            - version_number: Document version if present
            - effective_date: Document effective date

            **FIELD EXTRACTION (COMPREHENSIVE)**:
            - text_fields: All text input fields and values
            - numeric_fields: All numeric fields and values
            - date_fields: All date fields in various formats
            - checkbox_fields: Checkbox states (checked/unchecked)
            - radio_button_fields: Radio button selections
            - dropdown_selections: Selected options from dropdowns
            - signature_fields: Signature locations and names

            **SECTION ORGANIZATION**:
            - document_sections: Major document sections
            - subsections: Subsection organization
            - page_numbers: Page numbering if multi-page
            - cross_references: Internal document references

            **INSTRUCTIONAL ELEMENTS**:
            - instructions: Form completion instructions
            - help_text: Explanatory text and guidance
            - legal_disclaimers: Legal text and disclaimers
            - terms_conditions: Terms and conditions text
            - privacy_statements: Privacy policy text

            **ADMINISTRATIVE ELEMENTS**:
            - submission_instructions: How to submit document
            - contact_information: Contact details provided
            - office_use_only: Fields marked for official use
            - tracking_numbers: Reference or tracking numbers
            """

        # Custom fields addition
        custom_fields_protocol = ""
        if custom_fields:
            custom_fields_protocol = f"""
            **CUSTOM FIELD EXTRACTION (USER-SPECIFIED)**:
            The user has specifically requested extraction of these additional fields:
            {', '.join(custom_fields)}
            
            You MUST make every effort to locate and extract these fields even if they don't fit standard categories. Look for:
            - Exact text matches for these field names
            - Similar or related terms
            - Fields that might contain this information
            - Inferred or calculated values based on available data
            """

        # Output format specification
        output_format = """
        **CRITICAL: FAITHFUL DOCUMENT REPRODUCTION**
        You are creating a DIGITAL TWIN of this document. Every character, every field, every language must be reproduced with absolute fidelity. DO NOT improvise, translate, or alter ANYTHING.

        **CRITICAL JSON OUTPUT REQUIREMENTS**:
        Your response MUST be VALID JSON that can be parsed without errors.
        
        JSON FORMATTING RULES:
        - Use DOUBLE quotes for all strings and property names
        - NO trailing commas anywhere
        - NO unescaped quotes inside string values
        - NO single quotes - only double quotes
        - PROPERLY escape special characters with backslashes
        - Your entire response must be a single, complete JSON object
        - NO COMMENTS of any kind (no // comments, no /* */ comments)
        - NO placeholder text like "// ... other fields"
        - NEVER use "// ... " or similar comment syntax
        - EVERY field must be fully populated with actual data
        - NO shortcuts, abbreviations, or placeholder comments
        - COMPLETE the entire JSON structure with real extracted data
        
        **MANDATORY JSON OUTPUT FORMAT WITH SPATIAL INTELLIGENCE**:
        Your response MUST be a single, complete JSON object with this EXACT structure:

        {
            "document_type": "confirmed_document_type",
            "document_classification_confidence": 0.98,
            "extraction_confidence": 0.95,
            "processing_timestamp": "ISO_datetime",
            "document_dimensions": {
                "width": "document_width_pixels",
                "height": "document_height_pixels"
            },
            "spatial_layout": {
                "grid_system": {
                    "rows": "number_of_logical_rows",
                    "columns": "number_of_logical_columns"
                },
                "visual_zones": [
                    {
                        "zone_id": "header|body|footer|left_panel|right_panel",
                        "coordinates": [x1, y1, x2, y2],
                        "content_type": "text|image|logo|stamp",
                        "alignment": "left|center|right|justified"
                    }
                ]
            },
            "extracted_fields": {
                "field_name_1": {
                    "value": "extracted_value_1",
                    "confidence": 0.92,
                    "data_type": "string",
                    "extraction_method": "ocr",
                    "original_text": "raw_text_as_found_1",
                    "coordinates": [100, 200, 300, 220],
                    "dimensions": {"width": 200, "height": 20},
                    "excel_position": {"row": 1, "column": 1},
                    "alignment": "left",
                    "font_size": "12",
                    "language": "english",
                    "visual_group": "header_group",
                    "reading_order": "1",
                    "notes": "clear_text_extraction"
                },
                "field_name_2": {
                    "value": "extracted_value_2",
                    "confidence": 0.88,
                    "data_type": "string",
                    "extraction_method": "ocr",
                    "original_text": "raw_text_as_found_2",
                    "coordinates": [100, 240, 300, 260],
                    "dimensions": {"width": 200, "height": 20},
                    "excel_position": {"row": 2, "column": 1},
                    "alignment": "left",
                    "font_size": "12",
                    "language": "english",
                    "visual_group": "body_group",
                    "reading_order": "2",
                    "notes": "standard_field"
                }
            },
            "raw_text_complete": "ALL_EXTRACTED_TEXT_FROM_DOCUMENT",
            "text_regions": [
                {
                    "region_id": 1,
                    "text": "actual_text_from_region_1",
                    "coordinates": [0, 0, 500, 100],
                    "language": "english",
                    "confidence": 0.95
                },
                {
                    "region_id": 2,
                    "text": "actual_text_from_region_2",
                    "coordinates": [0, 100, 500, 200],
                    "language": "english",
                    "confidence": 0.90
                }
            ],
            "visual_elements": {
                "photos": ["description_of_photos"],
                "logos": ["description_of_logos"],
                "stamps": ["description_of_stamps"],
                "signatures": ["description_of_signatures"],
                "watermarks": ["description_of_watermarks"],
                "other_graphics": ["description_of_other_elements"]
            },
            "document_quality_assessment": {
                "overall_quality": "excellent|good|fair|poor",
                "text_clarity": "excellent|good|fair|poor", 
                "image_resolution": "high|medium|low",
                "lighting_conditions": "excellent|good|fair|poor",
                "rotation_angle": "degrees_if_rotated",
                "detected_languages": ["list_of_languages"],
                "ocr_challenges": ["list_of_difficulties"]
            },
            "extraction_metadata": {
                "total_fields_extracted": number,
                "high_confidence_fields": number,
                "medium_confidence_fields": number,
                "low_confidence_fields": number,
                "extraction_duration_estimate": "processing_time_estimate",
                "special_processing_notes": "any_important_notes",
                "recommended_human_verification": ["fields_needing_verification"]
            },
            "validation_results": {
                "date_format_consistency": true,
                "numeric_field_validation": true,
                "cross_field_consistency": true,
                "completeness_score": 0.95,
                "anomalies_detected": ["list_of_anomalies"]
            }
        }

        **CRITICAL REQUIREMENTS FOR SUCCESS**:
        1. Extract EVERY visible character - do not skip ANYTHING
        2. Process ALL languages present with equal precision
        3. Maintain original formatting and context
        4. Provide confidence scores based on OCR quality and certainty
        5. Include spatial information where possible
        6. Cross-validate all extracted information
        7. Flag any uncertainties or ambiguities
        8. Preserve cultural and linguistic nuances
        9. Handle mixed language documents expertly
        10. Apply document type expertise throughout extraction

        **ZERO TOLERANCE POLICY**:
        - NO missing text or fields
        - NO approximations or guesses without confidence scores
        - NO cultural insensitivity in processing
        - NO loss of information during extraction
        - NO incomplete analysis

        **FINAL JSON OUTPUT REQUIREMENTS - ABSOLUTELY CRITICAL**:
        
        RESPOND WITH NOTHING BUT COMPLETE VALID JSON. 
        
        DO NOT INCLUDE:
        - Any explanation before or after the JSON
        - Any markdown formatting (```json or ```)
        - Any comments inside the JSON (no // or /* */)
        - Any placeholder text like "... other fields"
        - Any incomplete structures
        - Any text outside the JSON object
        
        EXAMPLE OF WHAT NOT TO DO:
        ❌ "field_name": "value"    // ... other fields with similar structure
        ❌ "other_fields": "see above pattern"
        ❌ ```json { "data": "value" } ```
        
        EXAMPLE OF CORRECT OUTPUT:
        ✅ {"field_1": {"value": "actual_data", "confidence": 0.95}, "field_2": {"value": "more_data", "confidence": 0.88}}
        
        Your ENTIRE response must be parseable by json.loads() without ANY modifications.
        Start your response with { and end with }. Nothing else.
        
        RESPONSE LENGTH: Keep your response concise but complete. Avoid overly verbose descriptions.
        Focus on actual extracted values rather than long explanations.
        
        This is a MISSION-CRITICAL extraction. The extracted data will be used for official purposes.
        Your reputation as the world's best document forensics expert depends on PERFECT execution.
        """

        return base_forensic_protocol + specific_protocol + custom_fields_protocol + output_format


    
    


def convert_pdf_to_images(pdf_bytes: bytes) -> List[Image.Image]:
    """Convert PDF pages to images with high quality."""
    images = []
    try:
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page in pdf_document:
            pix = page.get_pixmap(dpi=300)  # Higher DPI for better OCR
            img_bytes = pix.tobytes("png")
            pil_image = Image.open(io.BytesIO(img_bytes))
            images.append(pil_image)
        return images
    except Exception as e:
        st.error(f"Failed to convert PDF: {e}")
        return []

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

def combine_multi_page_data(all_pages_data, filename):
    """Create metadata summary for multi-page document without mixing page data."""
    if not all_pages_data:
        return {}
    
    # Create a metadata summary without mixing the actual page data
    combined_data = {
        'document_type': all_pages_data[0].get('document_type', 'multi_page_document'),
        'document_pages': len(all_pages_data),
        'is_multi_page': len(all_pages_data) > 1,
        'filename': filename,
        'processing_timestamp': datetime.now().isoformat(),
        
        # Calculate overall statistics  
        'extraction_confidence': sum(float(p.get('extraction_confidence', 0)) if p.get('extraction_confidence') is not None else 0.0 for p in all_pages_data) / len(all_pages_data),
        'total_fields_extracted': sum(len(p.get('extracted_fields', {})) for p in all_pages_data),
        
        # Keep page-specific data separate for Excel processing
        'page_results': all_pages_data,
        
        # Create a summary extracted_fields for display purposes only (not mixing page data)
        'extracted_fields': {
            'document_summary': {
                'value': f"{len(all_pages_data)} pages processed with {sum(len(p.get('extracted_fields', {})) for p in all_pages_data)} total fields",
                'confidence': sum(float(p.get('extraction_confidence', 0)) if p.get('extraction_confidence') is not None else 0.0 for p in all_pages_data) / len(all_pages_data),
                'notes': 'Multi-page document summary - see individual page results in Excel'
            },
            'pages_processed': {
                'value': len(all_pages_data),
                'confidence': 1.0,
                'notes': 'Total number of pages processed'
            }
        },
        
        # Combine raw text from all pages with clear separation
        'raw_text_complete': '\n\n'.join([
            f"{'='*50}\nPAGE {page_idx + 1} CONTENT\n{'='*50}\n{page_data.get('raw_text_complete', '')}"
            for page_idx, page_data in enumerate(all_pages_data)
            if page_data.get('raw_text_complete', '').strip()
        ])
    }
    
    return combined_data

def create_spatial_layout_excel(processed_pages, filename, doc_type):
    """Create Excel file that recreates the EXACT visual layout of the document."""
    if not processed_pages:
        return create_single_document_excel({}, filename, doc_type)
    
    wb = Workbook()
    
    # Remove default sheet if we have multiple pages
    if len(processed_pages) > 1:
        wb.remove(wb.active)
    
    for page_idx, page_data in enumerate(processed_pages):
        # Create sheet for this page
        sheet_name = f"Page_{page_idx + 1}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Get spatial layout data
        spatial_layout = page_data.get('spatial_layout', {})
        extracted_fields = page_data.get('extracted_fields', {})
        
        # Set up Excel grid to match document layout
        grid_system = spatial_layout.get('grid_system', {})
        # Handle non-numeric values from Gemini
        try:
            max_rows = int(grid_system.get('rows', 50)) if str(grid_system.get('rows', '')).isdigit() else 50
            max_rows = max(1, max_rows)  # Ensure at least 1
        except (ValueError, TypeError):
            max_rows = 50
            
        try:
            max_cols = int(grid_system.get('columns', 20)) if str(grid_system.get('columns', '')).isdigit() else 20
            max_cols = max(1, max_cols)  # Ensure at least 1
        except (ValueError, TypeError):
            max_cols = 20
        
        # Initialize Excel sheet with proper dimensions
        for row in range(1, max_rows + 1):
            ws.row_dimensions[row].height = 20
        for col in range(1, max_cols + 1):
            col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
            ws.column_dimensions[col_letter].width = 15
        
        # Place fields based on their spatial coordinates
        for field_name, field_info in extracted_fields.items():
            if isinstance(field_info, dict):
                excel_pos = field_info.get('excel_position', {})
                # Handle non-numeric values from Gemini
                try:
                    row = int(excel_pos.get('row', 1)) if str(excel_pos.get('row', '')).isdigit() else 1
                    row = max(1, row)  # Ensure at least 1
                except (ValueError, TypeError):
                    row = 1
                    
                try:
                    col = int(excel_pos.get('column', 1)) if str(excel_pos.get('column', '')).isdigit() else 1
                    col = max(1, col)  # Ensure at least 1
                except (ValueError, TypeError):
                    col = 1
                
                # Ensure we don't exceed Excel limits
                if row <= max_rows and col <= max_cols:
                    cell = ws.cell(row=row, column=col)
                    cell.value = field_info.get('value', '')
                    
                    # Apply formatting based on spatial data
                    alignment = field_info.get('alignment', 'left')
                    try:
                        font_size = int(field_info.get('font_size', 11)) if str(field_info.get('font_size', '')).isdigit() else 11
                    except (ValueError, TypeError):
                        font_size = 11
                    
                    cell.alignment = Alignment(
                        horizontal=alignment,
                        vertical='center',
                        wrap_text=True
                    )
                    cell.font = Font(size=font_size)
                    
                    # Add border for visual separation
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
        
        # Add document metadata at the top
        ws.merge_cells('A1:D1')
        header_cell = ws['A1']
        header_cell.value = f"SPATIAL LAYOUT RECREATION - {filename} (Page {page_idx + 1})"
        header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get the actual extracted data to determine what columns we need
        formatted_page_data = format_extraction_results(page_data)
        
        if formatted_page_data:
            # Analyze the extracted data to determine column structure dynamically
            all_field_keys = set()
            for field_name, field_info in formatted_page_data.items():
                if isinstance(field_info, dict):
                    all_field_keys.update(field_info.keys())
            
            # Build column headers dynamically based on what's actually in the data
            columns = ['FIELD NAME', 'VALUE']
            col_letters = ['A', 'B']
            
            # Add any other unique keys found in the data (skip confidence and notes)
            for key in sorted(all_field_keys):
                if key not in ['value', 'confidence', 'notes'] and str(key).upper() not in columns:
                    columns.append(str(key).upper().replace('_', ' '))
                    col_letters.append(chr(ord(col_letters[-1]) + 1))
            
            # Set headers dynamically
            for i, (col_letter, header) in enumerate(zip(col_letters, columns)):
                ws[f'{col_letter}6'] = header
                cell = ws[f'{col_letter}6']
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row = 7
            
            # Add data dynamically based on the structure we discovered
            for field_name, field_info in formatted_page_data.items():
                # Field name
                ws[f'A{current_row}'] = field_name.replace('_', ' ').title()
                
                # Value
                ws[f'B{current_row}'] = field_info.get('value', '')
                
                # Add other columns dynamically (skip confidence and notes)
                for i in range(2, len(columns)):
                    col_letter = col_letters[i]
                    column_name = columns[i].lower().replace(' ', '_')
                    
                    if column_name in field_info:
                        value = field_info[column_name]
                        ws[f'{col_letter}{current_row}'] = str(value)
                
                current_row += 1
            
            # Auto-adjust column widths dynamically
            for i, col_letter in enumerate(col_letters):
                if i == 0:  # Field name
                    ws.column_dimensions[col_letter].width = 30
                elif i == 1:  # Value  
                    ws.column_dimensions[col_letter].width = 40
                else:  # Other columns
                    ws.column_dimensions[col_letter].width = 20
                    
        else:
            # No data case
            ws['A6'] = "FIELD NAME"
            ws['B6'] = "VALUE" 
            for col in ['A', 'B']:
                cell = ws[f'{col}6']
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            ws['A7'] = "No data extracted from this page"
            ws['A7'].font = Font(italic=True, color="808080")
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 40
    
    # If only one page, rename the sheet
    if len(processed_pages) == 1:
        wb.active.title = "Page_1"
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def create_multi_page_document_excel(all_pages_data, filename, doc_type):
    """Create Excel file for multi-page document with separate page blocks."""
    if not all_pages_data:
        return create_single_document_excel({}, filename, doc_type)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Page Document"
    
    # Document header with wider merge
    ws.merge_cells('A1:D1')
    header_cell = ws['A1']
    header_cell.value = f"MULTI-PAGE DOCUMENT EXTRACTION - {filename}"
    header_cell.font = Font(bold=True, size=16, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Processing info in a table format
    ws['A3'] = "Processing Date:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['C3'] = "Document Type:"
    ws['D3'] = doc_type.replace('_', ' ').title()
    ws['A4'] = "Total Pages:"
    ws['B4'] = len(all_pages_data)
    ws['C4'] = "Processing Mode:"
    ws['D4'] = "Multi-Page Analysis"
    
    # Style the info row
    for col in ['A', 'B', 'C', 'D']:
        for row in [3, 4]:
            cell = ws[f'{col}{row}']
            if col in ['A', 'C']:  # Labels
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    current_row = 6
    
    # Process each page as separate table blocks
    for page_idx, page_data in enumerate(all_pages_data):
        # Major page separator with full width
        ws.merge_cells(f'A{current_row}:D{current_row}')
        page_separator = ws[f'A{current_row}']
        page_separator.value = f"▓▓▓ PAGE {page_idx + 1} EXTRACTION RESULTS ▓▓▓"
        page_separator.font = Font(bold=True, size=14, color="FFFFFF")
        page_separator.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        page_separator.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Add page metadata if available
        if page_data.get('extraction_confidence'):
            ws.merge_cells(f'A{current_row}:D{current_row}')
            confidence_cell = ws[f'A{current_row}']
            confidence = page_data.get('extraction_confidence', 0)
            # Ensure confidence is a number
            try:
                confidence = float(confidence) if confidence is not None else 0.0
            except (ValueError, TypeError):
                confidence = 0.0
            field_count = len(page_data.get('extracted_fields', {}))
            confidence_cell.value = f"Page {page_idx + 1} Statistics: {field_count} fields extracted | Confidence: {confidence:.0%}"
            confidence_cell.font = Font(italic=True)
            confidence_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            confidence_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        
        # Page data table headers (4-column layout for better organization)
        ws[f'A{current_row}'] = "FIELD NAME"
        ws[f'B{current_row}'] = "EXTRACTED VALUE"
        ws[f'C{current_row}'] = "CONFIDENCE"
        ws[f'D{current_row}'] = "NOTES"
        
        # Style the table headers
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{current_row}']
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_row += 1
        
        # Add page data with enhanced formatting
        page_fields = page_data.get('extracted_fields', {})
        formatted_page_data = format_extraction_results({'extracted_fields': page_fields})
        
        if formatted_page_data:
            for field_name, field_info in formatted_page_data.items():
                # Field name
                ws[f'A{current_row}'] = field_name.replace('_', ' ').title()
                ws[f'A{current_row}'].font = Font(bold=True)
                
                # Field value
                value = field_info.get('value', '')
                ws[f'B{current_row}'] = str(value)[:200] + "..." if len(str(value)) > 200 else str(value)
                
                # Confidence score
                confidence = field_info.get('confidence', 0)
                ws[f'C{current_row}'] = f"{confidence:.1%}" if confidence > 0 else "N/A"
                
                # Color code confidence
                if confidence >= 0.8:
                    ws[f'C{current_row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif confidence >= 0.6:
                    ws[f'C{current_row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif confidence > 0:
                    ws[f'C{current_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Notes
                notes = field_info.get('notes', '')
                ws[f'D{current_row}'] = notes if notes else ""
                
                current_row += 1
        else:
            # No data found for this page
            ws.merge_cells(f'A{current_row}:D{current_row}')
            no_data_cell = ws[f'A{current_row}']
            no_data_cell.value = "No extractable data found on this page"
            no_data_cell.font = Font(italic=True, color="808080")
            no_data_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        
        # Add substantial spacing between pages (3 empty rows)
        current_row += 3
    
    # Auto-adjust column widths for better readability
    ws.column_dimensions['A'].width = 35  # Field names
    ws.column_dimensions['B'].width = 50  # Values
    ws.column_dimensions['C'].width = 12  # Confidence
    ws.column_dimensions['D'].width = 25  # Notes
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def create_single_document_excel(doc_data, filename, doc_type):
    """
    Creates an Excel report for a single document.
    
    Args:
        doc_data (dict): Extracted data in structured format
        filename (str): Original filename
        doc_type (str): Document type (e.g. 'invoice', 'id_card')
    
    Returns:
        bytes: Excel file in bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Document Data"
    
    # Document header
    ws.merge_cells('A1:B1')
    header_cell = ws['A1']
    header_cell.value = f"DOCUMENT EXTRACTION - {filename}"
    header_cell.font = Font(bold=True, size=14, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Processing info
    ws['A2'] = "Processing Date:"
    ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A3'] = "Document Type:"
    ws['B3'] = doc_type.replace('_', ' ').title()
    
    # Empty row
    current_row = 5
    
    # Field headers
    ws[f'A{current_row}'] = "FIELD NAME"
    ws[f'B{current_row}'] = "EXTRACTED VALUE"
    
    # Style headers
    for col in ['A', 'B']:
        cell = ws[f'{col}{current_row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row += 1
    
    # Add extracted data
    for field_name, field_info in doc_data.items():
        ws[f'A{current_row}'] = field_name.replace('_', ' ').title()
        ws[f'B{current_row}'] = field_info.get('value', '')
        current_row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def format_extraction_results(extracted_data: Dict) -> Dict:
    """Format extraction results for optimal display."""
    if not extracted_data:
        return {}
    
    formatted_results = {}
    extracted_fields = extracted_data.get('extracted_fields', {})
    
    for field_name, field_data in extracted_fields.items():
        if isinstance(field_data, dict):
            confidence = field_data.get('confidence', 0.0)
            # Ensure confidence is always a valid float
            if confidence is None:
                confidence = 0.0
            try:
                confidence = float(confidence)
            except (TypeError, ValueError):
                confidence = 0.0
            
            formatted_results[field_name] = {
                'value': field_data.get('value', ''),
                'confidence': confidence,
                'notes': field_data.get('notes', '')
            }
        else:
            # Handle simple field format
            formatted_results[field_name] = {
                'value': str(field_data),
                'confidence': 0.8,
                'notes': 'Legacy format'
            }
    
    return formatted_results

def create_comprehensive_report(extracted_data: Dict, doc_type: str) -> str:
    """Create a comprehensive analysis report."""
    report = f"""# DOCUMENT EXTRACTION REPORT

## Document Analysis Summary
- **Document Type**: {doc_type.replace('_', ' ').title()}
- **Processing Date**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- **Overall Confidence**: {float(extracted_data.get('extraction_confidence', 0)) if extracted_data.get('extraction_confidence') is not None else 0.0:.1%}
- **Classification Confidence**: {extracted_data.get('document_classification_confidence', 0):.1%}

## Quality Assessment
"""
    
    quality = extracted_data.get('document_quality_assessment', {})
    if quality:
        report += f"""- **Overall Quality**: {quality.get('overall_quality', 'Unknown')}
- **Text Clarity**: {quality.get('text_clarity', 'Unknown')}
- **Image Resolution**: {quality.get('image_resolution', 'Unknown')}
- **Languages Detected**: {', '.join(quality.get('detected_languages', []))}
"""

    report += "\n## Extracted Fields\n"
    
    extracted_fields = extracted_data.get('extracted_fields', {})
    for field_name, field_data in extracted_fields.items():
        if isinstance(field_data, dict):
            value = field_data.get('value', 'N/A')
            confidence = field_data.get('confidence', 0.0)
            notes = field_data.get('notes', '')
            
            # Ensure confidence is valid for formatting
            if confidence is None:
                confidence = 0.0
            try:
                confidence = float(confidence)
            except (TypeError, ValueError):
                confidence = 0.0
            
            report += f"- **{field_name.replace('_', ' ').title()}**: {value} (Confidence: {confidence:.1%})"
            if notes:
                report += f" - *{notes}*"
            report += "\n"
    
    # Visual elements
    visual_elements = extracted_data.get('visual_elements', {})
    if visual_elements and any(visual_elements.values()):
        report += "\n## Visual Elements Detected\n"
        for element_type, elements in visual_elements.items():
            if elements:
                report += f"- **{element_type.title()}**: {', '.join(elements)}\n"
    
    # Raw text
    raw_text = extracted_data.get('raw_text_complete', '')
    if raw_text:
        report += f"\n## Complete Raw Text Extraction\n```\n{raw_text}\n```\n"
    
    # Metadata
    metadata = extracted_data.get('extraction_metadata', {})
    if metadata:
        report += "\n## Extraction Metadata\n"
        for key, value in metadata.items():
            report += f"- **{key.replace('_', ' ').title()}**: {value}\n"
    
    # Validation results
    validation = extracted_data.get('validation_results', {})
    if validation:
        report += "\n## Validation Results\n"
        for key, value in validation.items():
            report += f"- **{key.replace('_', ' ').title()}**: {value}\n"
    
    return report

def main():
    """Main function for the advanced document extraction app."""
    st.title("🔬 Document Data Extractor")
    st.markdown("**ZERO-LOSS** document analysis with complete metadata extraction • Powered by Gemini API (Gemini 2.5 pro)")
    
    # Initialize session state
    if "forensic_processor" not in st.session_state:
        try:
            st.session_state.forensic_processor = AdvancedDocumentProcessor()
        except Exception as e:
            st.error("Failed to initialize Document Processor. Check configuration.")
            st.stop()
    
    if "forensic_results" not in st.session_state:
        st.session_state.forensic_results = None
    if "forensic_images" not in st.session_state:
        st.session_state.forensic_images = None
    if "batch_results" not in st.session_state:
        st.session_state.batch_results = []
    if "batch_processing_complete" not in st.session_state:
        st.session_state.batch_processing_complete = False
    
    # Sidebar configuration
    with st.sidebar:
        st.header("🔬 Analysis Settings")
        
        uploaded_files = st.file_uploader(
            "📤 Upload Documents for Batch Analysis",
            type=["jpg", "jpeg", "png", "pdf"],
            accept_multiple_files=True,
            help="CTRL+Click to select multiple files, or drag & drop multiple files at once"
        )
        
        st.markdown("---")
        
        # Document type selection
        doc_types = {
            "identity_card": "🆔 Identity Document (ID Cards, Passports)",
            "invoice": "🧾 Financial Document (Invoices, Receipts)", 
            "medical": "🏥 Medical Document (Reports, Prescriptions)",
            "certificate": "🎓 Certificate/Diploma",
            "form": "📝 Generic Form/Application"
        }
        
        selected_type = st.selectbox(
            "Document Type",
            options=list(doc_types.keys()),
            format_func=lambda x: doc_types[x],
            help="Select document type for specialized extraction protocols"
        )
        
        st.markdown("---")
        
        # Analysis intensity
        st.subheader("🎯 Analysis Intensity")
        intensity = st.select_slider(
            "Extraction Precision Level",
            options=["Standard", "Enhanced", "Forensic", "Ultimate"],
            value="Forensic",
            help="Higher levels extract more metadata but take longer"
        )
        
        # Custom fields
        st.subheader("🔍 Custom Field Extraction")
        custom_fields_input = st.text_area(
            "Additional fields to extract",
            placeholder="Employee ID\nSerial Number\nReference Code\nCustom Field",
            help="Specify additional fields beyond standard extraction"
        )
        
        custom_fields = [field.strip() for field in custom_fields_input.split('\n') if field.strip()]
        
        st.markdown("---")
        
        if st.button("🗑️ Clear Analysis Results", width='stretch'):
            st.session_state.forensic_results = None
            st.session_state.forensic_images = None
            st.session_state.batch_results = []
            st.session_state.batch_processing_complete = False
            st.rerun()
    
    # Main content area - BATCH PROCESSING VERSION
    if uploaded_files and not st.session_state.batch_processing_complete:
        st.subheader(f"📄 Batch Processing - {len(uploaded_files)} Document(s)")
        
        # Batch processing
        if st.button("🚀 PROCESS ALL DOCUMENTS", type="primary", width='stretch'):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Initialize batch results in session state
            if "batch_results" not in st.session_state:
                st.session_state.batch_results = []
            
            for i, uploaded_file in enumerate(uploaded_files):
                # Create a container for this document's results
                doc_container = st.container()
                
                # Process file (handle multi-page PDFs)
                try:
                    if uploaded_file.type == "application/pdf":
                        with st.spinner("Converting PDF..."):
                            pdf_bytes = uploaded_file.getvalue()
                            images = convert_pdf_to_images(pdf_bytes)
                        if not images:
                            st.error(f"Could not process PDF: {uploaded_file.name}")
                            continue
                    else:
                        images = [Image.open(uploaded_file)]
                    
                    # Process EACH page with extraction
                    all_pages_processed = []
                    combined_image = images[0]  # Use first page for display
                    
                    for page_idx, page_image in enumerate(images):
                        # Extract data from this page
                        with st.spinner(f"Processing {uploaded_file.name}..."):
                            page_data = st.session_state.forensic_processor.extract_document_with_forensic_precision(
                                page_image,
                                selected_type,
                                custom_fields
                            )
                        
                        if page_data:
                            # Add page metadata
                            page_data['page_number'] = page_idx + 1
                            page_data['total_pages'] = len(images)
                            page_data['page_image'] = page_image
                            
                            all_pages_processed.append(page_data)
                    
                    # Store the processed pages data
                    if all_pages_processed:
                        extracted_data = {
                            'document_type': selected_type,
                            'total_pages': len(images),
                            'processed_pages': all_pages_processed,
                            'filename': uploaded_file.name
                        }
                    else:
                        st.error(f"Could not process any page of {uploaded_file.name}")
                        continue
                    
                    # Generate unique key for this document result
                    result_key = f"batch_result_{i}_{uploaded_file.name}"
                    
                    # Store result in batch_results list with image
                    result = {
                        'filename': uploaded_file.name,
                        'data': extracted_data,
                        'doc_type': selected_type,
                        'image': combined_image,  # Store the first page image for display
                        'all_images': images,  # Store all page images
                        'excel_data': create_spatial_layout_excel(extracted_data.get('processed_pages', [extracted_data]), uploaded_file.name, selected_type)
                    }
                    st.session_state.batch_results.append(result)
                    
                    with doc_container:
                        col1, col2 = st.columns([1, 2])
                        
                        with col1:
                            # Display document image (first page)
                            if combined_image:
                                st.image(combined_image, caption=f"Preview: {uploaded_file.name} (Page 1/{len(images)})", use_container_width=True)
                            
                        with col2:
                            st.success(f"✅ {uploaded_file.name} processed successfully!")
                            
                            # Display processed pages data
                            if 'processed_pages' in extracted_data and extracted_data['processed_pages']:
                                st.markdown("### 📋 Extracted Data Preview")
                                
                                total_fields = 0
                                for page_idx, page_data in enumerate(extracted_data['processed_pages']):
                                    page_fields = format_extraction_results(page_data)
                                    total_fields += len(page_fields)
                                
                                st.info(f"Total: {total_fields} fields extracted from {len(extracted_data['processed_pages'])} pages")
                                
                                # Show preview from first page
                                first_page = extracted_data['processed_pages'][0]
                                first_page_fields = format_extraction_results(first_page)
                                
                                if first_page_fields:
                                    st.markdown("**Sample Fields (Page 1):**")
                                    # Show first few fields as preview
                                    preview_count = 0
                                    for field_name, field_info in first_page_fields.items():
                                        if preview_count >= 3:  # Limit preview
                                            break
                                        if isinstance(field_info, dict):
                                            st.text_input(
                                                label=f"Page 1: {field_name.replace('_', ' ').title()}",
                                                value=str(field_info.get('value', 'N/A'))[:40] + "..." if len(str(field_info.get('value', ''))) > 40 else str(field_info.get('value', 'N/A')),
                                                key=f"preview_{i}_{field_name}_{preview_count}",
                                                disabled=True
                                            )
                                            preview_count += 1
                                else:
                                    st.warning("No fields extracted")
                            else:
                                st.warning("No processed pages data found")
                        
                        # Show processing status instead of download button during batch processing
                        st.info(f"📋 Processing complete - Download available after batch finishes")
                        
                        # Update progress (ensure value is between 0 and 1)
                        progress_value = min(1.0, (i + 1) / len(uploaded_files))
                        progress_bar.progress(progress_value)
                
                except Exception as e:
                    st.error(f"❌ Error processing {uploaded_file.name}: {str(e)}")
                    
            # Mark batch processing as complete
            st.session_state.batch_processing_complete = True
            progress_bar.empty()  # Clear progress bar  
            status_text.empty()   # Clear status text
            st.success(f"🎉 BATCH PROCESSING COMPLETE! Processed {len(st.session_state.batch_results)} documents")
            time.sleep(2)  # Show success message briefly
            st.rerun()
    
    # Show batch results if processing is complete
    elif st.session_state.batch_processing_complete and st.session_state.batch_results:
        st.subheader(f"📊 Batch Processing Results - {len(st.session_state.batch_results)} Documents")
        
        # Action buttons
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("🆕 Process New Batch", width='stretch'):
                st.session_state.batch_results = []
                st.session_state.batch_processing_complete = False
                st.rerun()
        
        with col2:
            # Create ZIP with all documents
            def create_batch_zip():
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for idx, result in enumerate(st.session_state.batch_results):
                        formatted_data = format_extraction_results(result['data'])
                        if formatted_data:
                            # Create Excel for this document
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill
                            
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Document Data"
                            
                            # Header info
                            ws['A1'] = f"Document: {result['filename']}"
                            ws['A2'] = f"Type: {result['doc_type']}"
                            ws['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                            
                            # Field headers
                            ws['A5'] = "Field Name"
                            ws['B5'] = "Extracted Value"
                            
                            for cell in [ws['A5'], ws['B5']]:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            
                            # Add data
                            row = 6
                            for field_name, field_info in formatted_data.items():
                                ws[f'A{row}'] = field_name.replace('_', ' ').title()
                                ws[f'B{row}'] = str(field_info.get('value', ''))
                                row += 1
                            
                            # Adjust columns
                            ws.column_dimensions['A'].width = 30
                            ws.column_dimensions['B'].width = 50
                            
                            # Save to buffer
                            excel_buffer = io.BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            # Add to ZIP
                            clean_name = result['filename'].replace('.', '_')
                            excel_filename = f"extracted_{clean_name}.xlsx"
                            zip_file.writestr(excel_filename, excel_buffer.getvalue())
                
                zip_buffer.seek(0)
                return zip_buffer.getvalue()
            
            # Download all as ZIP button with unique key
            zip_data = create_batch_zip()
            zip_key = f"download_zip_{len(st.session_state.batch_results)}_{int(time.time())}"
            st.download_button(
                "📦 Download All (ZIP)",
                zip_data,
                f"batch_extractions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                "application/zip",
                key=zip_key,
                type="primary",
                use_container_width=True
            )
        
        with col3:
            if st.button("🗑️ Clear Results", width='stretch'):
                st.session_state.batch_results = []
                st.session_state.batch_processing_complete = False
                st.rerun()
        
        st.markdown("---")
        
        # Display all batch results
        for idx, result in enumerate(st.session_state.batch_results):
            with st.expander(f"📄 {result['filename']}", expanded=(idx == 0)):
                # Create layout with image and data side by side
                col_img, col_data = st.columns([1, 2])
                
                with col_img:
                    # Display document image if available
                    if 'image' in result and result['image'] is not None:
                        st.image(result['image'], caption=f"Original: {result['filename']}", use_container_width=True)
                    else:
                        st.info("Image not available")
                
                with col_data:
                    # Extract data for display - HANDLE PROCESSED PAGES
                    extracted_data = result['data']
                    
                    # Check if this is processed pages data
                    if 'processed_pages' in extracted_data and extracted_data['processed_pages']:
                        processed_pages = extracted_data['processed_pages']
                        total_fields = sum(len(format_extraction_results(page)) for page in processed_pages)
                        
                        # Show metrics
                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            st.metric("Fields Extracted", total_fields)
                        with col_m2:
                            avg_confidence = sum(float(page.get('extraction_confidence', 0)) if page.get('extraction_confidence') is not None else 0.0 for page in processed_pages) / len(processed_pages) if processed_pages else 0
                            st.metric("Confidence", f"{avg_confidence:.0%}")
                        with col_m3:
                            st.metric("Pages", len(processed_pages))
                        
                        # Display extracted fields from each page
                        st.markdown("### 📋 Extracted Data")
                        
                        for page_idx, page_data in enumerate(processed_pages):
                            with st.expander(f"📄 Page {page_idx + 1}", expanded=(page_idx == 0)):
                                page_fields = format_extraction_results(page_data)
                                
                                if page_fields:
                                    field_items = list(page_fields.items())
                                    num_cols = 2
                                    for i in range(0, len(field_items), num_cols):
                                        cols = st.columns(num_cols)
                                        for j in range(num_cols):
                                            if i + j < len(field_items):
                                                field_name, field_info = field_items[i+j]
                                                with cols[j]:
                                                    st.metric(
                                                        label=field_name.replace('_', ' ').title(),
                                                        value=str(field_info.get('value', 'N/A'))
                                                    )
                                else:
                                    st.info("No data from this page")
                    
                    else:
                        # Fallback to old logic for single pages
                        formatted_data = format_extraction_results(extracted_data)
                        
                        # Show metrics
                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            field_count = len(formatted_data) if formatted_data else 0
                            st.metric("Fields Extracted", field_count)
                        with col_m2:
                            confidence = extracted_data.get('extraction_confidence', 0)
                            try:
                                confidence = float(confidence) if confidence is not None else 0.0
                            except (ValueError, TypeError):
                                confidence = 0.0
                            st.metric("Confidence", f"{confidence:.0%}")
                        with col_m3:
                            quality = extracted_data.get('document_quality_assessment', {}).get('overall_quality', 'N/A')
                            st.metric("Quality", quality.title())
                        
                        # Display extracted fields
                        if formatted_data:
                            st.markdown("### 📋 Extracted Data")
                            
                            field_items = list(formatted_data.items())
                            num_cols = 2
                            for i in range(0, len(field_items), num_cols):
                                cols = st.columns(num_cols)
                                for j in range(num_cols):
                                    if i + j < len(field_items):
                                        field_name, field_info = field_items[i+j]
                                        with cols[j]:
                                            st.metric(
                                                label=field_name.replace('_', ' ').title(),
                                                value=field_info.get('value', 'N/A')
                                            )
                
                # Create Excel download for this document
                def create_single_excel(doc_data, filename, doc_type):
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill, Alignment
                        
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Document Data"
                        
                        # Header
                        ws['A1'] = f"Document: {filename}"
                        ws['A2'] = f"Type: {doc_type}"
                        ws['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                        
                        # Fields
                        ws['A5'] = "Field Name"
                        ws['B5'] = "Extracted Value"
                        
                        # Style header
                        for col in ['A', 'B']:
                            cell = ws[f'{col}5']
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        
                        # Add data
                        row = 6
                        for field_name, field_info in doc_data.items():
                            ws[f'A{row}'] = field_name.replace('_', ' ').title()
                            ws[f'B{row}'] = str(field_info.get('value', ''))
                            row += 1
                        
                        # Adjust column widths
                        ws.column_dimensions['A'].width = 30
                        ws.column_dimensions['B'].width = 50
                        
                        # Save to buffer
                        buffer = io.BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)
                        return buffer.getvalue()
                    except Exception as e:
                        st.error(f"Excel creation failed: {e}")
                        return None
                
                # Download button with unique key to prevent app restart
                # Use the result's excel_data that was created during processing
                clean_name = result['filename'].replace('.', '_')
                # Create stable unique key using file content hash instead of time
                import hashlib
                content_hash = hashlib.md5(result['filename'].encode()).hexdigest()[:8]
                unique_key = f"download_single_{idx}_{content_hash}"
                st.download_button(
                    f"📊 Download Excel - {result['filename']}",
                    result['excel_data'],
                    f"extracted_{clean_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=unique_key,
                    use_container_width=True
                )
            
            # Display results if available
            if st.session_state.forensic_results:
                results = st.session_state.forensic_results
                extracted_data = results['data']
                
                if extracted_data:
                    # Analysis metrics
                    col_a, col_b, col_c = st.columns(3)
                    
                    with col_a:
                        confidence = extracted_data.get('extraction_confidence', 0)
                        try:
                            confidence = float(confidence) if confidence is not None else 0.0
                        except (ValueError, TypeError):
                            confidence = 0.0
                        st.metric("Extraction Confidence", f"{confidence:.1%}")
                    
                    with col_b:
                        field_count = len(extracted_data.get('extracted_fields', {}))
                        st.metric("Fields Extracted", field_count)
                    
                    with col_c:
                        quality = extracted_data.get('document_quality_assessment', {})
                        overall_quality = quality.get('overall_quality', 'unknown')
                        st.metric("Document Quality", overall_quality.title())
                    
                    # Extracted fields with advanced display
                    st.subheader("📋 Extracted Data Fields")
                    
                    formatted_results = format_extraction_results(extracted_data)
                    
                    if formatted_results:
                        # Create tabs for different field categories
                        field_tabs = st.tabs(["📝 All Fields", "⭐ High Confidence", "⚠️ Needs Review"])
                        
                        with field_tabs[0]:  # All fields
                            edited_fields = {}
                            for field_name, field_info in formatted_results.items():
                                col_field, col_conf, col_notes = st.columns([2, 1, 1])
                                
                                with col_field:
                                    current_value = field_info.get('value', '')
                                    edited_value = st.text_input(
                                        f"**{field_name.replace('_', ' ').title()}**",
                                        value=str(current_value) if current_value else '',
                                        key=f"field_{field_name}",
                                        help=field_info.get('notes', '')
                                    )
                                    edited_fields[field_name] = edited_value
                                
                                with col_conf:
                                    confidence = field_info.get('confidence', 0.0)
                                    if confidence is None:
                                        confidence = 0.0
                                    confidence = float(confidence) if confidence is not None else 0.0
                                    if confidence >= 0.9:
                                        st.success(f"{confidence:.1%}")
                                    elif confidence >= 0.7:
                                        st.warning(f"{confidence:.1%}")
                                    else:
                                        st.error(f"{confidence:.1%}")
                                
                                with col_notes:
                                    notes = field_info.get('notes', '')
                                    if notes:
                                        st.caption(f"ℹ️ {notes}")
                        
                        with field_tabs[1]:  # High confidence fields
                            high_conf_fields = {k: v for k, v in formatted_results.items() 
                                              if v.get('confidence', 0) >= 0.8}
                            if high_conf_fields:
                                for field_name, field_info in high_conf_fields.items():
                                    st.write(f"✅ **{field_name.replace('_', ' ').title()}**: {field_info.get('value', '')}")
                            else:
                                st.info("No high-confidence fields found. Consider higher quality image.")
                        
                        with field_tabs[2]:  # Needs review
                            low_conf_fields = {k: v for k, v in formatted_results.items() 
                                             if v.get('confidence', 0) < 0.8}
                            if low_conf_fields:
                                for field_name, field_info in low_conf_fields.items():
                                    st.write(f"⚠️ **{field_name.replace('_', ' ').title()}**: {field_info.get('value', '')} "
                                           f"({field_info.get('confidence', 0):.1%})")
                            else:
                                st.success("All extracted fields have high confidence!")
                    
                    # Export options
                    st.subheader("💾 Export Analysis Results")
                    
                    col_export1, col_export2, col_export3 = st.columns(3)
                    
                    with col_export1:
                        # Complete JSON export
                        json_data = json.dumps(extracted_data, indent=2, ensure_ascii=False)
                        st.download_button(
                            "📋 Complete JSON Data",
                            json_data,
                            f"forensic_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                            "application/json"
                        )
                    
                    with col_export2:
                        # Comprehensive report
                        report = create_comprehensive_report(extracted_data, results['doc_type'])
                        st.download_button(
                            "📄  Report",
                            report,
                            f"forensic_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
                            "text/markdown"
                        )
                    
                    with col_export3:
                        # Field data only (CSV-like)
                        field_data = []
                        for field_name, field_info in formatted_results.items():
                            field_data.append(f"{field_name},{field_info.get('value', '')},{field_info.get('confidence', 0)}")
                        
                        csv_content = "Field,Value,Confidence\n" + "\n".join(field_data)
                        st.download_button(
                            "📊 Field Data CSV",
                            csv_content,
                            f"extracted_fields_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            "text/csv"
                        )
                    
                    # Advanced analysis sections
                    with st.expander("🔍 Complete Raw Text Extraction", expanded=False):
                        raw_text = extracted_data.get('raw_text_complete', '')
                        if raw_text:
                            st.text_area("All extracted text", raw_text, height=300, disabled=True)
                        else:
                            st.info("Raw text not available in analysis results.")
                    
                    with st.expander("🎯 Visual Elements Analysis", expanded=False):
                        visual_elements = extracted_data.get('visual_elements', {})
                        if visual_elements:
                            for element_type, elements in visual_elements.items():
                                if elements:
                                    st.write(f"**{element_type.title()}**: {', '.join(elements)}")
                        else:
                            st.info("Visual elements analysis not available.")
                    
                    with st.expander("⚙️ Technical Analysis Details", expanded=False):
                        # Quality assessment
                        quality = extracted_data.get('document_quality_assessment', {})
                        if quality:
                            st.subheader("Quality Assessment")
                            for key, value in quality.items():
                                st.write(f"**{key.replace('_', ' ').title()}**: {value}")
                        
                        # Metadata
                        metadata = extracted_data.get('extraction_metadata', {})
                        if metadata:
                            st.subheader("Processing Metadata")
                            for key, value in metadata.items():
                                st.write(f"**{key.replace('_', ' ').title()}**: {value}")
                        
                        # Validation
                        validation = extracted_data.get('validation_results', {})
                        if validation:
                            st.subheader("Validation Results")
                            for key, value in validation.items():
                                st.write(f"**{key.replace('_', ' ').title()}**: {value}")
                
                else:
                    st.error("❌  analysis failed. Please check document quality and try again.")

if __name__ == "__main__":
    main()