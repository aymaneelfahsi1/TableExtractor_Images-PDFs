import os
import sys
import logging
import json
import re
from dataclasses import dataclass
from typing import List, Dict, Optional

# --- Core Libraries ---
import streamlit as st
from PIL import Image
import google.generativeai as genai

# --- PDF Handling ---
try:
    import fitz  # PyMuPDF
except ImportError:
    st.error("PyMuPDF is not installed. Please run 'pip install PyMuPDF' to handle PDF files.")
    sys.exit(1)

# --- Excel Handling ---
try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import pandas as pd
except ImportError:
    st.error("Required libraries for Excel are not installed. Please run 'pip install openpyxl pandas'")
    sys.exit(1)

# --- App Configuration ---
st.set_page_config(page_title="Intelligent Table Extractor", page_icon="📄", layout="wide")

# --- Logging ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - [%(levelname)s] - %(message)s')

# --- Gemini Configuration & Backend Logic ---
@dataclass
class TableProcessorConfig:
    """Configuration for the advanced table processing agent."""
    # This assumes you have a secrets.toml file with your GOOGLE_API_KEY
    GOOGLE_API_KEY: str = st.secrets.get("GOOGLE_API_KEY", "")
    GEMINI_MODEL: str = "gemini-2.5-pro" # Using the latest recommended model
    OUTPUT_DIR: str = "extracted_tables"

class SingleShotTableProcessor:
    """
    The backend engine for the Streamlit app, powered by a self-correcting AI agent
    with semantic and hierarchical structure understanding.
    """
    def __init__(self):
        self.config = TableProcessorConfig()
        if not self.config.GOOGLE_API_KEY:
            st.error("Your GOOGLE_API_KEY is not set! Please add it to your Streamlit secrets.")
            st.stop()
        self._initialize_gemini()
        os.makedirs(self.config.OUTPUT_DIR, exist_ok=True)

    def _initialize_gemini(self):
        try:
            genai.configure(api_key=self.config.GOOGLE_API_KEY)
            self.model = genai.GenerativeModel(self.config.GEMINI_MODEL)
            logging.info("Gemini initialized successfully.")
        except Exception as e:
            st.error(f"Failed to initialize Gemini. The API key may be invalid. Error: {e}")
            raise

    def extract_all_tables_in_one_shot(self, img: Image.Image, table_hint: int) -> List[Dict]:
        """
        Processes an entire image using a robust, single-shot prompt.
        """
        try:
            master_prompt = f"""
            You are a meticulous document forensics expert. Your primary goal is to perform a structural analysis of the document image, identifying all tables and their complete hierarchical context, and structure them into a flawless JSON list.

            A user has provided a hint that they expect to find around **{table_hint}** table(s). Use this as a guide, but your own expert analysis is paramount.

            YOUR FORENSIC ANALYSIS & SELF-CORRECTION PROTOCOL:
            1.  **IDENTIFY & ISOLATE TABLES:** Scan the entire image to identify every distinct table. Ignore all non-tabular data.
            2.  **HIERARCHICAL HEADER ANALYSIS (CRITICAL):** For each table, you must analyze the header structure for multiple levels. A header is a "parent header" if it visibly spans across multiple columns below it. The columns below a parent are its "subheaders."
            3.  **SEMANTIC DISAMBIGUATION (CRITICAL):** You must ensure all final column headers are unique.
                -   If simple headers are duplicated (e.g., two 'DATE' columns), rename them based on context (e.g., 'DATE_Start', 'DATE_End').
                -   Sub-headers under different parents are distinct. Your `column_structure` output should preserve this hierarchy.
            4.  **DATA & SUMMARY EXTRACTION:** Extract all data rows and identify any final summary/total rows.
            5.  **FAITHFULNESS TO STRUCTURE:** If a cell is visually empty, it MUST be represented as an empty string `""`. DO NOT invent data.
            6.  **FORMULA VALIDATION:** Internally verify summary calculations. If a value is a sum of the column above, create a `formula_instructions` entry with `column_index` and `formula_type`.

            REQUIRED JSON OUTPUT STRUCTURE:
            Your entire response MUST be a single JSON list `[...]`. Each table object MUST follow this precise hierarchical schema.

            GENERIC STRUCTURAL EXAMPLE:
            {{
              "table_id": 1,
              "table_title": "Title of the Table (if found)",
              "column_structure": [
                {{ "name": "Simple Header", "subheaders": [] }},
                {{ "name": "Parent Header A", "subheaders": [{{ "name": "Sub-Header 1" }}, {{ "name": "Sub-Header 2" }}] }}
              ],
              "data_rows": [ ... ],
              "summary_row": [ ... ],
              "formula_instructions": [ {{ "column_index": 1, "formula_type": "SUM_ABOVE" }} ]
            }}

            Respond ONLY with the raw JSON list.
            """
            response = self.model.generate_content([master_prompt, img])
            cleaned_text = response.text.strip().replace("```json", "").replace("```", "")
            if not cleaned_text:
                st.warning("The Analyst (Gemini) returned an empty response.")
                return []
            all_tables_instructions = json.loads(cleaned_text)
            st.info(f"Analyst successfully found and processed {len(all_tables_instructions)} table(s).")
            return all_tables_instructions
        except Exception as e:
            st.error(f"An error occurred during the Gemini analysis process: {e}")
            logging.error(f"Gemini analysis error. Response text was: {response.text if 'response' in locals() else 'N/A'}")
            return []

    def create_excel_file(self, instructions: Dict, output_path: str, table_num: int) -> Optional[str]:
        """
        Builds a flawless Excel file from the hierarchical instruction format, creating merged cells.
        """
        if not instructions or "column_structure" not in instructions:
            return None
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            gemini_title = instructions.get("table_title")
            safe_title = gemini_title if gemini_title else f"Table_{table_num}"
            sanitized_title = re.sub(r'[\\/*?:"<>|]', "_", safe_title)
            ws.title = sanitized_title[:31]

            header_font = Font(bold=True, size=12)
            cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')
            
            column_structure = instructions.get("column_structure", [])
            has_subheaders = any(h.get("subheaders") for h in column_structure)
            header_rows = 2 if has_subheaders else 1
            current_col = 1
            flat_headers = flatten_hierarchical_headers(column_structure)

            for header in column_structure:
                cell = ws.cell(row=1, column=current_col, value=header.get("name"))
                cell.font, cell.border, cell.alignment = header_font, cell_border, center_align
                subheaders = header.get("subheaders", [])
                if subheaders and isinstance(subheaders, list) and len(subheaders) > 0:
                    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(subheaders) - 1)
                    for i, sub in enumerate(subheaders):
                        sub_cell = ws.cell(row=2, column=current_col + i, value=sub.get("name"))
                        sub_cell.font, sub_cell.border = header_font, cell_border
                    current_col += len(subheaders)
                else:
                    if has_subheaders: ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
                    current_col += 1
            
            current_row = header_rows + 1
            data_start_row = current_row
            data_rows = instructions.get("data_rows", [])
            for row_data in data_rows:
                for c_idx, cell_data in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=c_idx)
                    cell.border = cell_border
                    if cell_data is None or str(cell_data).strip() == "": cell.value = None
                    else:
                        try: cell.value = float(str(cell_data).replace(",", "").strip())
                        except (ValueError, TypeError): cell.value = cell_data
                current_row += 1
            data_end_row = current_row - 1

            summary_row_data = instructions.get("summary_row", [])
            if summary_row_data:
                for c_idx, cell_data in enumerate(summary_row_data, 1):
                    cell = ws.cell(row=current_row, column=c_idx)
                    cell.value = cell_data
                    cell.border, cell.font = cell_border, header_font
            
            for instruction in instructions.get("formula_instructions", []):
                target_col_index = instruction.get("column_index")
                formula_type = instruction.get("formula_type")
                if target_col_index is not None and formula_type == "SUM_ABOVE":
                    target_col = target_col_index + 1
                    col_letter = get_column_letter(target_col)
                    formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                    target_cell = ws.cell(row=current_row, column=target_col)
                    target_cell.value, target_cell.font = formula, header_font
                else:
                    logging.warning(f"Skipping malformed formula instruction: {instruction}")
            
            for i in range(1, len(flat_headers) + 1):
                try:
                    max_len = max(len(str(cell.value or "")) for cell in ws[get_column_letter(i)])
                    ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
                except ValueError: # Handle empty columns
                    continue
                    
            wb.save(output_path)
            return output_path
        except Exception as e:
            st.error(f"Failed to create Excel file {output_path}: {e}")
            return None

# --- Helper Functions & UI ---
def convert_pdf_to_images(pdf_bytes: bytes) -> List[Image.Image]:
    """
    Converts each page of a PDF file into a PIL Image object.
    """
    try:
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
        # --- ❗ FIX WAS APPLIED HERE ❗ ---
        # The 'Page' object from fitz does not have a direct '.width' or '.height'.
        # You must access the page's rectangle '.rect' first.
        # It's also good practice to convert the dimensions to integers.
        return [
            Image.frombytes("RGB", [int(p.rect.width), int(p.rect.height)], p.get_pixmap().samples)
            for p in pdf_document
        ]
    except Exception as e:
        st.error(f"Failed to convert PDF: {e}")
        logging.error(f"PyMuPDF conversion failed: {e}")
        return []

def flatten_hierarchical_headers(column_structure: List[Dict]) -> List[str]:
    """
    Converts the nested header structure from JSON into a single flat list for pandas.
    """
    flat_headers = []
    for header in column_structure:
        parent_name = header.get("name", "")
        subheaders = header.get("subheaders", [])
        if subheaders and isinstance(subheaders, list) and len(subheaders) > 0:
            for sub in subheaders:
                sub_name = sub.get("name", "")
                flat_headers.append(f"{parent_name}_{sub_name}" if parent_name else sub_name)
        else:
            flat_headers.append(parent_name)
    return flat_headers

def main():
    """Main function to run the Streamlit app."""
    st.title("📄 Table Extractor using Gemini Pro")
    st.markdown("An AI agent that performs a forensic analysis of documents to extract tables.")

    # --- Initialize Session State ---
    if "processor" not in st.session_state:
        try:
            st.session_state.processor = SingleShotTableProcessor()
        except Exception as e:
            st.error("Failed to initialize the AI Processor. Is your API key set correctly in Streamlit secrets?")
            st.stop()
    if "results" not in st.session_state:
        st.session_state.results = None
    if "images_to_display" not in st.session_state:
        st.session_state.images_to_display = None
    if "page_configs" not in st.session_state:
        st.session_state.page_configs = {}

    # --- Sidebar for Configuration ---
    with st.sidebar:
        st.header("⚙️ Configuration")
        st.info("API Key is managed via Streamlit Secrets.")
        st.markdown("---")
        
        uploaded_file = st.file_uploader("Upload a new Image or PDF to start", type=["jpg", "jpeg", "png", "pdf"])
        
        # --- DYNAMIC UI FOR PDF PAGE SELECTION ---
        selected_pages_indices = []
        if uploaded_file and uploaded_file.type == "application/pdf":
            try:
                with st.spinner("Reading PDF..."):
                    pdf_bytes = uploaded_file.getvalue()
                    pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                    page_options = [f"Page {i+1}" for i in range(len(pdf_doc))]
            
                selected_pages_str = st.multiselect(
                    "Select pages to process",
                    options=page_options,
                    default=page_options[0] if page_options else None
                )
                selected_pages_indices = [int(p.split(" ")[1]) - 1 for p in selected_pages_str]
            except Exception as e:
                st.error(f"Could not read the PDF file. It might be corrupted. Error: {e}")
                st.stop()

        global_table_hint = st.slider("Default Table Hint (per page)", 1, 10, 1, 1, help="This will be the default hint for all selected pages.")
        
        if st.sidebar.button("Clear Results", use_container_width=True):
            st.session_state.results, st.session_state.images_to_display, st.session_state.page_configs = None, None, {}
            st.rerun()

    # --- MAIN CONTENT AREA ---
    if uploaded_file:
        if uploaded_file.type == "application/pdf":
            if selected_pages_indices:
                st.subheader("⚙️ Per-Page Configuration")
                st.markdown("Set a specific table hint for each page you selected.")
                page_configs = {}
                for page_index in selected_pages_indices:
                    col1, col2 = st.columns([1, 3])
                    with col1:
                        st.markdown(f"**Page {page_index + 1}**")
                    with col2:
                        hint = st.number_input(
                            f"Table hint for page {page_index+1}",
                            min_value=1, max_value=10,
                            value=st.session_state.page_configs.get(page_index, global_table_hint),
                            step=1, key=f"hint_{page_index}"
                        )
                        page_configs[page_index] = hint
                st.session_state.page_configs = page_configs
                
                if st.button("🚀 Start Analysis on Selected Pages", use_container_width=True, type="primary"):
                    process_file(uploaded_file, selected_pages_indices)
        else: # Handle Image files
            st.session_state.page_configs = {0: global_table_hint} # For single image, page index is 0
            if st.button("🚀 Start Analysis", use_container_width=True, type="primary"):
                process_file(uploaded_file, [0]) # Process page at index 0

    if st.session_state.results:
        display_results()

def process_file(uploaded_file, selected_pages_indices: List[int]):
    """Handles the main processing logic when the button is clicked."""
    st.session_state.results, st.session_state.images_to_display = None, None
    
    with st.spinner("Preparing document..."):
        file_bytes = uploaded_file.getvalue()
        all_images = convert_pdf_to_images(file_bytes) if uploaded_file.type == "application/pdf" else [Image.open(uploaded_file)]
    
    if not all_images:
        st.error("Could not extract any images from the uploaded file.")
        return

    images_to_process = [(idx, all_images[idx]) for idx in selected_pages_indices if idx < len(all_images)]
    st.session_state.images_to_display = images_to_process
    
    results_by_page = {}
    with st.spinner("AI is analyzing... This may take a moment."):
        for page_index, img in images_to_process:
            page_hint = st.session_state.page_configs.get(page_index, 1)
            st.write(f"Analyzing Page {page_index + 1} with a hint of {page_hint} table(s)...")
            instructions = st.session_state.processor.extract_all_tables_in_one_shot(img, page_hint)
            results_by_page[page_index] = instructions
    
    st.session_state.results = results_by_page
    st.rerun()

def display_results():
    """Renders the results stored in the session state."""
    st.markdown("---")
    st.header("📊 Analysis Results")
    total_tables_found = 0
    
    for page_index, img in st.session_state.images_to_display:
        page_num = page_index + 1
        page_results = st.session_state.results.get(page_index)

        with st.container(border=True):
            st.subheader(f"📄 Page {page_num}")
            col1, col2 = st.columns([1, 2])
            col1.image(img, caption=f"Analyzed Page {page_num}", use_column_width=True)

            with col2:
                if not page_results:
                    st.warning(f"No tables were found on Page {page_num}.")
                    continue

                for j, table_instructions in enumerate(page_results):
                    total_tables_found += 1
                    title = table_instructions.get("table_title") or f"Table {j + 1} on Page {page_num}"
                    with st.expander(f"**{title}**", expanded=True):
                        
                        column_structure = table_instructions.get("column_structure", [])
                        data_rows = table_instructions.get("data_rows", [])
                        summary_row = table_instructions.get("summary_row", [])
                        flat_headers = flatten_hierarchical_headers(column_structure)

                        if flat_headers and (data_rows or summary_row):
                            try:
                                preview_data = data_rows + ([summary_row] if summary_row else [])
                                df = pd.DataFrame(preview_data, columns=flat_headers)
                                st.dataframe(df)
                            except Exception as e:
                                st.error(f"Could not display preview for '{title}': {e}")
                                st.json(table_instructions)

                        output_filename = f"Page_{page_num}_Table_{j + 1}.xlsx"
                        output_path = os.path.join(st.session_state.processor.config.OUTPUT_DIR, output_filename)
                        saved_path = st.session_state.processor.create_excel_file(table_instructions, output_path, total_tables_found)

                        if saved_path:
                            with open(saved_path, "rb") as file:
                                st.download_button(
                                    f"📥 Download {output_filename}",
                                    file,
                                    output_filename,
                                    key=f"dl_{page_num}_{j}"
                                )
                        else:
                            st.error(f"Could not generate Excel file for '{title}'.")

    st.markdown("---")
    st.success(f"**Analysis complete! Found a total of {total_tables_found} tables across the selected pages.**")

if __name__ == "__main__":
    main()

